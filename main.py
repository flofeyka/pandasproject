import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys
import time

class try_toconnect():
    try:
        global wb
        global my_sheet
        global df
        df = pd.read_excel('./DataBase.xlsx')
        wb = openpyxl.load_workbook('./DataBase.xlsx')
        my_sheet = wb.active
    except:
        print('Failed to connect to the DB')
        time.sleep(5)
        sys.exit()
def main():
    print ('Sorting the table: 1. Descending\n2. Ascending\n3. The current database\n4. The number of values \n5. Alphabetical name\n6. At first...\n7. Back')
    value = int(input("Enter value: "))
    if value == 1:
        print('Sorting how?\n1. By the year of entry \n2. In descending order of age\n3. In descending order of age and year of entry\n4. Back')
        sorting_dis = int(input('Enter value: '))
        if sorting_dis == 1:
            print(df.sort_values(["Year_to_join"], ascending=False))
        elif sorting_dis == 2:
            print(df.sort_values(["Age"]), ascending=False)
        elif sorting_dis == 3:
            print(df.sort_values(["Age","Year_to_join"], ascending=[False, False]))
        else:
            return main()
    elif value == 2:
        print("Sorting how?\n1. By the year of entry\n2. In ascending order of age\n3. In ascending order of age and year of entry\n4. Back")
        sorting_asc = int(input('Enter value: '))
        if sorting_asc == 1:
            print(df.sort_values(["Year_to_join"]))
        elif sorting_asc == 2:
            print(df.sort_values(["Age"]))
        elif sorting_asc == 3:
            print(df.sort_values(["Age","Year_to_join"]))
        else:
            return main()
    elif value == 3:
        print(df)
        return main()
    elif value == 4:
        print('1. The number of owners of one smartphone brand\n2. The number of any single roads\n3. The number of people of the same sex\n4. The number of one-year-olds\n5. The number of owners of one OS\n6. The number of one-year-olds joining the conversation \n7. The number of people of similar nationality \n8. The number of clients of one bank\n9. Back')
        value2 = int(input('Enter value: '))
        if value2 == 1:
            print(df["Brand of phone"].value_counts())
        elif value2 == 2:
            print(df["City"].value_counts())
        elif value2 == 3:
            print(df["Sex"].value_counts())
        elif value2 == 4:
            print(df["Age"].value_counts())
        elif value2 == 5:
            print(df["OS"].value_counts())
        elif value2 == 6:
            print(df["Year_to_join"].value_counts())
        elif value2 == 7:
            print(df["National"].value_counts())
        elif value2 == 8:
            print(df["Bank"].value_counts())
        else:
            return main()
    elif value == 5:
        print("1. From A to Z\n2. From Z to A\n3. Back")
        value3 = int(input("Enter value: "))
        if value3 == 1:
            print(df.sort_values(["Name"]))
        elif value3 == 2:
            print(df.sort_values(["Name"], ascending=False))
        else:
            return main()

    elif value == 6:
        def value6():
            print("First what:\n1. First, the residents of Russia/Kazakhstan\n2. First Android/iOS\n3. First high/low\n4. Back")
            num = int(input("Enter value: "))
            if num == 1:
                print('1. First, the residents of Russia\n2. First, residents of Kazakhstan')
                numc = int(input("Enter value: "))
                if numc == 1:
                    print(df.sort_values(["Country"], ascending=False))
                elif numc == 2:
                    print(df.sort_values(["Country"]))
                else:
                    return value6()
            elif num == 2:
                print('1. Android first\n2. First iOS\n3. Back')
                numo = int(input("Enter value: "))
                if numo == 1:
                    print(df.sort_values(["OS"]))
                elif numo == 2:
                    print(df.sort_values(["OS"], ascending=False))
                else: 
                    return value6()
            elif num == 3:
                print('1. First high \n2. First low\n3. Back')
                numh = int(input("Enter value: "))
                if numh == 1:
                    print(df.sort_values(["Height"]))
                elif numh == 2:
                    print(df.sort_values(["Height"], ascending=False))
                else:
                    return value6()
            else:
                return main()
        value6()
    else:
        return

def redactor():
    print('What do you want to do?\n1.Editing/adding data\n2.Deleting the line \n3.Back')
    main = int(input("Enter value: "))
    if main == 1:
        print (pd.read_excel("./DataBase.xlsx"))
        
        Name = str(input('Enter your first and last name: '))
        Age = int(input("How old are you: "))
        City = str(input("What's your city: "))
        Country = str(input("What's your country: "))
        Sex = str(input("What's your gendet: "))
        Height = int(input("How much is your height: "))
        Grade = int(input("Enter the school class (if you graduated from school, enter 0): "))
        OS = str(input("Enter your operating system on your phone(iOS/Android): "))
        Brand_of_phone = str(input("Enter the brand of your smartphone:"))
        National = str(input("Enter your nationality:"))
        Birth = str(input("Enter your date of birth(example: 01.01.2021):"))
        Bank = str(input("Enter your main bank:"))
        Year_to_join = int(input("Enter the year you entered the conversation (example: 2020):"))


        dbredactor = int(input("Specify the line number (In case you want to edit the data, just specify the line with the existing data):")) + 2

        print('Do you really want to edit line', inti-2, '\n1.Yes\n2.No')
        sure = int(input("Enter value: "))
        if sure == 1:
            c1 = my_sheet.cell(row = inti, column = 1)
            c1.value = Name
            c2 = my_sheet.cell(row= inti, column = 2)
            c2.value = Age
            c3 = my_sheet.cell(row=inti, column=3)
            c3.value = City
            c4 = my_sheet.cell(row=inti,column=4)
            c4.value = Country
            c5 = my_sheet.cell(row=inti,column=5)
            c5.value = Sex
            c6 = my_sheet.cell(row=inti,column=7)
            c6.value = Height
            c7 = my_sheet.cell(row=inti, column=8)
            c7.value = Grade
            c8 = my_sheet.cell(row=inti, column=9)
            c8.value = OS
            c9 = my_sheet.cell(row=inti,column=10)
            c9.value = Brand_of_phone
            c10 = my_sheet.cell(row=inti, column=11)
            c10.value = National
            c11 = my_sheet.cell(row=inti, column=12)
            c11.value = Birth
            c12 = my_sheet.cell(row=inti, column=13)
            c12.value = Bank
            c13 = my_sheet.cell(row=inti, column=14)
            c13.value = Year_to_join
            c14 = my_sheet.cell(row=inti, column=6)
            c14.value = 'Non admin'
            wb.save("./DataBase.xlsx")
            print(pd.read_excel('./DataBase.xlsx'))
        else:
            return redactor()

    elif main == 2:
        print(df)
        roww = int(input("Enter the row number:")) + 2
        print('Are you sure you want to delete row ',roww,'?\n1.Yes\n2.No')
        sure2 = int(input('Enter value: '))
        if sure2 == 1:
            my_sheet.delete_rows(roww)
            wb.save('./DataBase.xlsx')

            print(pd.read_excel('./DataBase.xlsx'))
        else:
            return redactor()
    else:
        return

        
class somain:
    wb = openpyxl.load_workbook('DataBase.xlsx')
    wb.save("./DataBase.xlsx")
    df = pd.read_excel('./DataBase.xlsx')
    print('Welcome to the console sorter and DB and DT editor.\What do you want to do?\n1.Sort \n2.Edit')
    a = int(input("Enter value: "))
    if a==1:
        main()
    elif a==2:
        redactor()
    else:
        print("Wrong value")

while True:
    somain()
    try_toconnect()
